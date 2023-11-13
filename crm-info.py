# oreas catalogue info-grabber
# started 2023/10/02 ZH

versionNum = "v0.1.0"
versionDate = "2023/11/13"

import os
import sys
import re
import pandas as pd
import numpy as np
import chemparse
import json
from collections import Counter
from functools import cache, wraps
from time import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from analysismethodprefs import analysis_method_prefs_dict as PREFS
from analysismethodprefs import supercrm_list as SUPERCRMS


# for sorting analysis method, USE: 4-Acid Digestion >
# chris has doc: Merging.xlsx G:\.shortcut-targets-by-id\1w2nUsja1tidZ-QYTuemO6DzCaclAmIlm\PXRFS\12. Certified Reference Material\5_CRM Master Spreadsheet\OREAS Sorting
# get Principle cert vals and which are superCRMs from copying search


def timing(f):
    @wraps(f)
    def wrap(*args, **kw):
        ts = time()
        result = f(*args, **kw)
        te = time()
        print("func:%r args:[%r, %r] took: %2.4f sec" % (f.__name__, args, kw, te - ts))
        return result

    return wrap


class Crm:
    def __init__(
        self,
        crm_id: str,
        crm_group: str,
        crm_type: str,
        crm_matrix: str,
        crm_mineralisation: str,
        crm_status: str,
    ):
        self.id = crm_id
        self.group = crm_group
        self.type = crm_type
        self.matrix = crm_matrix
        self.mineralisation = crm_mineralisation
        self.status = crm_status
        self.supercrm = "yes" if isSuperCRM(crm_id) else "no"
        self.url = ""
        if crm_id.startswith(
            "OREAS "
        ):  # if oreas CRM, format ID to comply with chris' preferred formatting i.e. 'OREAS 100a' -> 'OREAS0100a' and generate link
            self.id = format_oreas_crm_id(crm_id)
            self.url = generate_oreas_crm_url(crm_id)
        self.chemistry_ppm = {}
        self.chemistry_wtpercent = {}

        # print(f'Initialising CRM... {self.id=} {self.group=}, {self.type=}, {self.matrix=}, {self.mineralisation=}, {self.status=}, {self.supercrm=}')

    def addChemistry(
        self,
        chem_formula: str,
        chem_concentration,
        chem_unit: str,
        chem_analysis_method: str,
        chem_is_indicative=False,
    ):
        # check formula - attempt conversion
        new_element = getFirstElementOfCompound(chem_formula)
        conv_factor = compoundToElementConversionFactor(chem_formula)
        if conv_factor == 1:
            new_element = chem_formula  # if cannot do stoich, (i.e. invalid symbols), then just resort to compound conc.

        # GET PPM CONCENTRATION
        # check unit - convert if necessary ( can just multiply by conv_factor from compound conv. fact)
        conv_factor_ppm = conv_factor * getUnitConversionFactor(
            units_from=chem_unit, units_to="ppm"
        )
        new_concentration_ppm = chem_concentration * conv_factor_ppm

        # GET % CONCENTRATION
        # check unit - convert if necessary ( can just multiply by conv_factor from compound conv. fact)
        conv_factor_wtpercent = conv_factor * getUnitConversionFactor(
            units_from=chem_unit, units_to="%"
        )
        new_concentration_wtpercent = chem_concentration * conv_factor_wtpercent

        new_element_with_units_ppm = f"{new_element}(PPM)"
        new_element_with_units_wtpercent = f"{new_element}(%)"

        # add to BOTH chem dicts
        if new_element_with_units_ppm not in self.chemistry_ppm.keys():
            # print(f'adding {new_element} to chemistry dict for {self.id}...')
            self.chemistry_ppm[new_element_with_units_ppm] = {}
            self.chemistry_wtpercent[new_element_with_units_wtpercent] = {}

        # rounding values to 3 dec places for ppm (lowest dp will be 1ppb), and 7 for % (lowest dp wll be 1ppb)
        if chem_is_indicative:
            self.chemistry_ppm[new_element_with_units_ppm][
                chem_analysis_method
            ] = f"{np.round(new_concentration_ppm,decimals=3)} INDICATIVE"
            self.chemistry_wtpercent[new_element_with_units_wtpercent][
                chem_analysis_method
            ] = f"{np.round(new_concentration_wtpercent,decimals=7)} INDICATIVE"
        else:
            self.chemistry_ppm[new_element_with_units_ppm][
                chem_analysis_method
            ] = np.round(new_concentration_ppm, decimals=3)
            self.chemistry_wtpercent[new_element_with_units_wtpercent][
                chem_analysis_method
            ] = np.round(new_concentration_wtpercent, decimals=7)

        # no longer rounding conc values per chris' request
        # self.chemistry[new_element][chem_analysis_method] = (
        #     np.round(new_concentration, decimals=0)
        #     if new_concentration > 20
        #     else np.round(new_concentration, decimals=3)
        # )
        # rounds new ppm values to 3 decimal places (first place is 1ppb)
        # if len(str(new_concentration)) > 8:
        #     print(f'{self.id}: {new_element} conc: {new_concentration} rounded to {np.round(new_concentration,decimals=0) if new_concentration > 20 else np.round(new_concentration,decimals=3)}')


def isSuperCRM(crm_id: str):
    if crm_id in SUPERCRMS:
        return True
    else:
        return False


def getUnitConversionFactor(units_from: str, units_to: str):
    """Returns the factor to multiply by to convert from units_from to units_to"""
    if units_from == "wt.%" or units_from == "wt %":
        units_from = "%"
    if units_to == "wt.%" or units_to == "wt %":
        units_to = "%"

    if units_from == "Unity":
        return 1

    if units_from == units_to:
        return 1
    elif units_from == "ppm" and units_to == "ppb":
        return 1000
    elif units_from == "ppm" and units_to == "%":
        return 1 / 10000
    elif units_from == "ppb" and units_to == "ppm":
        return 1 / 1000
    elif units_from == "ppb" and units_to == "%":
        return 1 / 10000000
    elif units_from == "%" and units_to == "ppm":
        return 10000
    elif units_from == "%" and units_to == "ppb":
        return 10000000
    else:
        raise Exception(
            f"Invalid unit conversion requested! {units_from=}, {units_to=}"
        )


@cache
def format_oreas_crm_id(crm_id: str):
    """given an oreas crm id directly from the catalogue, return chris' requested format for the name (e.g. 'OREAS 45f' -> 'OREAS0045f')"""
    # Define regex pattern to match desired formatting from chris
    pattern = r"^OREAS (\d+)"

    # use re.sub to apply pattern to crm id strings, and replace space?
    formatted_string = re.sub(
        pattern, lambda match: f"OREAS {match.group(1).zfill(4)}", crm_id
    ).replace(" ", "")

    return formatted_string


@cache
def generate_oreas_crm_url(crm_id: str):
    """returns a url for oreas' website for the page for a CRM, given the crm ID from the catalogue. (e.g. 'OREAS 45f' -> 'https://www.oreas.com/crm/oreas-45f/')"""
    # as of 2023/11/10, formatting for oreas site CRM details page is: https://www.oreas.com/crm/oreas-45f/
    # i.e. replace spaces with hyphens
    adjusted_crm_id = crm_id.replace(" ", "-")
    return f'=Hyperlink("https://www.oreas.com/crm/{adjusted_crm_id}/")'


@cache
def getFirstElementOfCompound(compound: str):
    eoi_predicted = compound[0]
    strposition = 1
    cont = True
    while cont and (strposition < len(compound)):
        if compound[strposition].isupper():
            # print(f'{compound[strposition]} is upper!')
            cont = False
        elif compound[strposition].isnumeric():
            # print(f'{compound[strposition]} is number!')
            cont = False
        else:
            eoi_predicted += compound[strposition]
        strposition += 1

    return eoi_predicted


@cache
def getFirstElementOfNameAndCompound(nameandcompound: str):
    """for oreas catalogue 'Element' column. e.g. input: 'Iron(III) oxide, Fe2O3' -> 'Fe'."""
    if ", " not in nameandcompound:
        print(f"no comma in {nameandcompound}!!")
        raise Exception
    compound = nameandcompound.split(", ")[1]

    element_symbol_matchstrs = [
        "H",
        "He",
        "Li",
        "Be",
        "B",
        "C",
        "N",
        "O",
        "F",
        "Ne",
        "Na",
        "Mg",
        "Al",
        "Si",
        "P",
        "S",
        "Cl",
        "Ar",
        "K",
        "Ca",
        "Sc",
        "Ti",
        "V",
        "Cr",
        "Mn",
        "Fe",
        "Co",
        "Ni",
        "Cu",
        "Zn",
        "Ga",
        "Ge",
        "As",
        "Se",
        "Br",
        "Kr",
        "Rb",
        "Sr",
        "Y",
        "Zr",
        "Nb",
        "Mo",
        "Tc",
        "Ru",
        "Rh",
        "Pd",
        "Ag",
        "Cd",
        "In",
        "Sn",
        "Sb",
        "Te",
        "I",
        "Xe",
        "Cs",
        "Ba",
        "La",
        "Ce",
        "Pr",
        "Nd",
        "Pm",
        "Sm",
        "Eu",
        "Gd",
        "Tb",
        "Dy",
        "Ho",
        "Er",
        "Tm",
        "Yb",
        "Lu",
        "Hf",
        "Ta",
        "W",
        "Re",
        "Os",
        "Ir",
        "Pt",
        "Au",
        "Hg",
        "Tl",
        "Pb",
        "Bi",
        "Po",
        "At",
        "Rn",
        "Fr",
        "Ra",
        "Ac",
        "Th",
        "Pa",
        "U",
        "Np",
        "Pu",
        "Am",
        "Cm",
        "Bk",
        "Cf",
        "Es",
        "Fm",
        "Md",
        "No",
        "Lr",
        "Rf",
        "Db",
        "Sg",
        "Bh",
        "Hs",
        "Mt",
        "Ds",
        "Rg",
        "Cn",
        "Nh",
        "Fl",
        "Mc",
        "Lv",
        "Ts",
        "Og",
    ]
    elementsymbol = getFirstElementOfCompound(compound)
    if elementsymbol in element_symbol_matchstrs:
        return elementsymbol
    else:
        return compound


def compoundToElementConversionFactor(compound: str, element_of_interest: str = ""):
    """Returns the conversion factor to apply to a compound concentration to get the
    concentration of the element of interest (eoi), given the compound and eoi as strs.
    e.g. given 'Al2O3' and 'Al', returns 0.529251.
    If no eoi is given, will assume the first element symbol found in the compound str is the eoi.
    """

    def shouldAttemptStoichConversion(stoich_dict: dict):
        element_symbol_matchstrs = [
            "H",
            "He",
            "Li",
            "Be",
            "B",
            "C",
            "N",
            "O",
            "F",
            "Ne",
            "Na",
            "Mg",
            "Al",
            "Si",
            "P",
            "S",
            "Cl",
            "Ar",
            "K",
            "Ca",
            "Sc",
            "Ti",
            "V",
            "Cr",
            "Mn",
            "Fe",
            "Co",
            "Ni",
            "Cu",
            "Zn",
            "Ga",
            "Ge",
            "As",
            "Se",
            "Br",
            "Kr",
            "Rb",
            "Sr",
            "Y",
            "Zr",
            "Nb",
            "Mo",
            "Tc",
            "Ru",
            "Rh",
            "Pd",
            "Ag",
            "Cd",
            "In",
            "Sn",
            "Sb",
            "Te",
            "I",
            "Xe",
            "Cs",
            "Ba",
            "La",
            "Ce",
            "Pr",
            "Nd",
            "Pm",
            "Sm",
            "Eu",
            "Gd",
            "Tb",
            "Dy",
            "Ho",
            "Er",
            "Tm",
            "Yb",
            "Lu",
            "Hf",
            "Ta",
            "W",
            "Re",
            "Os",
            "Ir",
            "Pt",
            "Au",
            "Hg",
            "Tl",
            "Pb",
            "Bi",
            "Po",
            "At",
            "Rn",
            "Fr",
            "Ra",
            "Ac",
            "Th",
            "Pa",
            "U",
            "Np",
            "Pu",
            "Am",
            "Cm",
            "Bk",
            "Cf",
            "Es",
            "Fm",
            "Md",
            "No",
            "Lr",
            "Rf",
            "Db",
            "Sg",
            "Bh",
            "Hs",
            "Mt",
            "Ds",
            "Rg",
            "Cn",
            "Nh",
            "Fl",
            "Mc",
            "Lv",
            "Ts",
            "Og",
        ]
        if not stoich_dict:  # if stoich dict is empty
            return False
        for key in stoich_dict.keys():
            if key not in element_symbol_matchstrs:
                return False
        return True

    compound_stoich_dict = {}
    compound_stoich_dict = chemparse.parse_formula(
        compound
    )  # returns e.g. {'Al': 2, 'O': 3} from 'Al2O3'
    if not shouldAttemptStoichConversion(compound_stoich_dict):
        return 1

    if element_of_interest == "":
        element_of_interest = getFirstElementOfCompound(compound)

    compound_mass = 0
    eoi_mass_single = 0
    eoi_mass = 0

    masses = {
        "H": 1.00794,
        "He": 4.002602,
        "Li": 6.941,
        "Be": 9.012182,
        "B": 10.811,
        "C": 12.0107,
        "N": 14.0067,
        "O": 15.9994,
        "F": 18.9984032,
        "Ne": 20.1797,
        "Na": 22.98976928,
        "Mg": 24.305,
        "Al": 26.9815386,
        "Si": 28.0855,
        "P": 30.973762,
        "S": 32.065,
        "Cl": 35.453,
        "Ar": 39.948,
        "K": 39.0983,
        "Ca": 40.078,
        "Sc": 44.955912,
        "Ti": 47.867,
        "V": 50.9415,
        "Cr": 51.9961,
        "Mn": 54.938045,
        "Fe": 55.845,
        "Co": 58.933195,
        "Ni": 58.6934,
        "Cu": 63.546,
        "Zn": 65.409,
        "Ga": 69.723,
        "Ge": 72.64,
        "As": 74.9216,
        "Se": 78.96,
        "Br": 79.904,
        "Kr": 83.798,
        "Rb": 85.4678,
        "Sr": 87.62,
        "Y": 88.90585,
        "Zr": 91.224,
        "Nb": 92.90638,
        "Mo": 95.94,
        "Tc": 98.9063,
        "Ru": 101.07,
        "Rh": 102.9055,
        "Pd": 106.42,
        "Ag": 107.8682,
        "Cd": 112.411,
        "In": 114.818,
        "Sn": 118.71,
        "Sb": 121.760,
        "Te": 127.6,
        "I": 126.90447,
        "Xe": 131.293,
        "Cs": 132.9054519,
        "Ba": 137.327,
        "La": 138.90547,
        "Ce": 140.116,
        "Pr": 140.90465,
        "Nd": 144.242,
        "Pm": 146.9151,
        "Sm": 150.36,
        "Eu": 151.964,
        "Gd": 157.25,
        "Tb": 158.92535,
        "Dy": 162.5,
        "Ho": 164.93032,
        "Er": 167.259,
        "Tm": 168.93421,
        "Yb": 173.04,
        "Lu": 174.967,
        "Hf": 178.49,
        "Ta": 180.9479,
        "W": 183.84,
        "Re": 186.207,
        "Os": 190.23,
        "Ir": 192.217,
        "Pt": 195.084,
        "Au": 196.966569,
        "Hg": 200.59,
        "Tl": 204.3833,
        "Pb": 207.2,
        "Bi": 208.9804,
        "Po": 208.9824,
        "At": 209.9871,
        "Rn": 222.0176,
        "Fr": 223.0197,
        "Ra": 226.0254,
        "Ac": 227.0278,
        "Th": 232.03806,
        "Pa": 231.03588,
        "U": 238.02891,
        "Np": 237.0482,
        "Pu": 244.0642,
        "Am": 243.0614,
        "Cm": 247.0703,
        "Bk": 247.0703,
        "Cf": 251.0796,
        "Es": 252.0829,
        "Fm": 257.0951,
        "Md": 258.0951,
        "No": 259.1009,
        "Lr": 262,
        "Rf": 267,
        "Db": 268,
        "Sg": 271,
        "Bh": 270,
        "Hs": 269,
        "Mt": 278,
        "Ds": 281,
        "Rg": 281,
        "Cn": 285,
        "Nh": 284,
        "Fl": 289,
        "Mc": 289,
        "Lv": 292,
        "Ts": 294,
        "Og": 294,
    }

    try:
        eoi_mass_single = masses[element_of_interest]
    except KeyError:
        print(
            f"Error: Supplied Element of Interest ({element_of_interest}) for compound ({compound}) not found in molecular mass dictionary"
        )
        return 1

    # Calculate actual EOI mass in case of multiple stoich of EOI in compound (e.g. Al2O3)
    try:
        eoi_mass = (
            eoi_mass_single * compound_stoich_dict[element_of_interest]
        )  # e.g. 26.9815386 * 2 = 53.9630772
    except KeyError:
        print(
            f"Error: Supplied Element of Interest ({element_of_interest}) not found in compound ({compound})"
        )
        return 1

    for element, quantity in compound_stoich_dict.items():
        try:
            compound_mass += masses[element] * quantity
        except KeyError:
            (
                f"Error: Element ({element}) in Compound ({compound}) not found in molecular mass dictionary"
            )
            return 1

    return eoi_mass / compound_mass


def countAnalysisMethodsForElement(catdf: pd.DataFrame, element: str):
    """pass oreas catalogue dataframe and element of interest. recv counter object for analysis methods for that element."""
    copydf = catdf.copy()
    print(copydf)
    # copydf.set_index('Element', inplace=True)
    # copydf = copydf.filter(like=element,axis=0)
    copydf = copydf[copydf["Element Symbol"] == element]
    print(copydf)
    method_counter = Counter(copydf["Analysis Method"].tolist())
    return method_counter


@timing
def main():
    # open oreas catalogue as csv, convert to dataframe
    catalogue_path = "oreas-catalogue-2023-11-03.csv"
    cat_df = pd.read_csv(catalogue_path)

    # add element symbol only AND superCRM columns
    cat_df["Element Symbol"] = cat_df["Element"].apply(getFirstElementOfNameAndCompound)
    cat_df["SuperCRM"] = cat_df["CRM ID"].apply(isSuperCRM)
    # print(cat_df)
    # cat_df.to_csv('output.csv')
    # print(countAnalysisMethodsForElement(cat_df,'U'))

    # initially process CRMs from catalogue into Crm class objects
    crm_ids_seen = set([])
    crms = []
    for i in cat_df.index:
        id = cat_df["CRM ID"][i]
        # if crm id not seen before, then make new instance of CRM class
        if id not in crm_ids_seen:
            if crm_ids_seen:  # if ids seen list is NOT empty
                crms.append(currentcrm)
                # append currentCRM to list unless it's the first crm (list will be empty)
            crm_ids_seen.add(id)
            currentcrm = Crm(
                crm_id=id,
                crm_group=cat_df["CRM Group"][i],
                crm_type=cat_df["CRM Type"][i],
                crm_matrix=cat_df["Matrix"][i],
                crm_mineralisation=cat_df["Mineralisation Style"][i],
                crm_status=cat_df["Status"][i],
            )
        # add data to crm
        formula = cat_df["Element"][i]
        if ", " in formula:
            formula = formula.split(", ")[1]
        # check if value is INDICATIVE, (also e.g.<2ppm) - if so, DO NOT USE
        if cat_df["1SD"][i] == "IND":
            currentcrm.addChemistry(
                chem_formula=formula,
                chem_concentration=cat_df["Certified Value"][i],
                chem_unit=cat_df["Unit"][i],
                chem_analysis_method=cat_df["Analysis Method"][i],
                chem_is_indicative=True,
            )
        else:
            currentcrm.addChemistry(
                chem_formula=formula,
                chem_concentration=cat_df["Certified Value"][i],
                chem_unit=cat_df["Unit"][i],
                chem_analysis_method=cat_df["Analysis Method"][i],
                chem_is_indicative=False,
            )

    crms.append(currentcrm)  # for the last on the list!

    crms_total_amount = len(crm_ids_seen)
    print(f"Total of {crms_total_amount} CRMs found in catalogue.")

    all_data_df = pd.DataFrame(
        columns=[
            "ID",
            "Group",
            "Type",
            "Matrix",
            "Mineralisation Style",
            "SuperCRM",
            "Status",
            "Catalogue File Name",
            "URL",
            "Ag(PPM)",
            "Ag(%)",
            "Ag Method",
            "Al(PPM)",
            "Al(%)",
            "Al Method",
            "As(PPM)",
            "As(%)",
            "As Method",
            "Au(PPM)",
            "Au(%)",
            "Au Method",
            "B(PPM)",
            "B(%)",
            "B Method",
            "Ba(PPM)",
            "Ba(%)",
            "Ba Method",
            "Be(PPM)",
            "Be(%)",
            "Be Method",
            "Bi(PPM)",
            "Bi(%)",
            "Bi Method",
            "Ca(PPM)",
            "Ca(%)",
            "Ca Method",
            "Cd(PPM)",
            "Cd(%)",
            "Cd Method",
            "Ce(PPM)",
            "Ce(%)",
            "Ce Method",
            "Cl(PPM)",
            "Cl(%)",
            "Cl Method",
            "Co(PPM)",
            "Co(%)",
            "Co Method",
            "Cr(PPM)",
            "Cr(%)",
            "Cr Method",
            "Cs(PPM)",
            "Cs(%)",
            "Cs Method",
            "Cu(PPM)",
            "Cu(%)",
            "Cu Method",
            "Dy(PPM)",
            "Dy(%)",
            "Dy Method",
            "Er(PPM)",
            "Er(%)",
            "Er Method",
            "Eu(PPM)",
            "Eu(%)",
            "Eu Method",
            "Fe(PPM)",
            "Fe(%)",
            "Fe Method",
            "Ga(PPM)",
            "Ga(%)",
            "Ga Method",
            "Gd(PPM)",
            "Gd(%)",
            "Gd Method",
            "Ge(PPM)",
            "Ge(%)",
            "Ge Method",
            "Hf(PPM)",
            "Hf(%)",
            "Hf Method",
            "Hg(PPM)",
            "Hg(%)",
            "Hg Method",
            "Ho(PPM)",
            "Ho(%)",
            "Ho Method",
            "In(PPM)",
            "In(%)",
            "In Method",
            "Ir(PPM)",
            "Ir(%)",
            "Ir Method",
            "K(PPM)",
            "K(%)",
            "K Method",
            "La(PPM)",
            "La(%)",
            "La Method",
            "Li(PPM)",
            "Li(%)",
            "Li Method",
            "Lu(PPM)",
            "Lu(%)",
            "Lu Method",
            "Mg(PPM)",
            "Mg(%)",
            "Mg Method",
            "Mn(PPM)",
            "Mn(%)",
            "Mn Method",
            "Mo(PPM)",
            "Mo(%)",
            "Mo Method",
            "Na(PPM)",
            "Na(%)",
            "Na Method",
            "Nb(PPM)",
            "Nb(%)",
            "Nb Method",
            "Nd(PPM)",
            "Nd(%)",
            "Nd Method",
            "Ni(PPM)",
            "Ni(%)",
            "Ni Method",
            "P(PPM)",
            "P(%)",
            "P Method",
            "Pb(PPM)",
            "Pb(%)",
            "Pb Method",
            "Pd(PPM)",
            "Pd(%)",
            "Pd Method",
            "Pr(PPM)",
            "Pr(%)",
            "Pr Method",
            "Pt(PPM)",
            "Pt(%)",
            "Pt Method",
            "Rb(PPM)",
            "Rb(%)",
            "Rb Method",
            "Re(PPM)",
            "Re(%)",
            "Re Method",
            "Rh(PPM)",
            "Rh(%)",
            "Rh Method",
            "Ru(PPM)",
            "Ru(%)",
            "Ru Method",
            "S(PPM)",
            "S(%)",
            "S Method",
            "Sb(PPM)",
            "Sb(%)",
            "Sb Method",
            "Sc(PPM)",
            "Sc(%)",
            "Sc Method",
            "Se(PPM)",
            "Se(%)",
            "Se Method",
            "Si(PPM)",
            "Si(%)",
            "Si Method",
            "Sm(PPM)",
            "Sm(%)",
            "Sm Method",
            "Sn(PPM)",
            "Sn(%)",
            "Sn Method",
            "Sr(PPM)",
            "Sr(%)",
            "Sr Method",
            "Ta(PPM)",
            "Ta(%)",
            "Ta Method",
            "Tb(PPM)",
            "Tb(%)",
            "Tb Method",
            "Te(PPM)",
            "Te(%)",
            "Te Method",
            "Th(PPM)",
            "Th(%)",
            "Th Method",
            "Ti(PPM)",
            "Ti(%)",
            "Ti Method",
            "Tl(PPM)",
            "Tl(%)",
            "Tl Method",
            "Tm(PPM)",
            "Tm(%)",
            "Tm Method",
            "U(PPM)",
            "U(%)",
            "U Method",
            "V(PPM)",
            "V(%)",
            "V Method",
            "W(PPM)",
            "W(%)",
            "W Method",
            "Y(PPM)",
            "Y(%)",
            "Y Method",
            "Yb(PPM)",
            "Yb(%)",
            "Yb Method",
            "Zn(PPM)",
            "Zn(%)",
            "Zn Method",
            "Zr(PPM)",
            "Zr(%)",
            "Zr Method",
        ]
    )
    conc = 0
    statusi = 0
    for crm in crms:
        statusi += 1
        print(f"Processing CRM Data... [{str(statusi).zfill(3)}/{crms_total_amount}]")
        crm_row_dict = {}
        crm_row_dict["ID"] = crm.id
        crm_row_dict["Group"] = crm.group
        crm_row_dict["Type"] = crm.type
        crm_row_dict["Matrix"] = crm.matrix
        crm_row_dict["Mineralisation Style"] = crm.mineralisation
        crm_row_dict["SuperCRM"] = crm.supercrm
        crm_row_dict["Status"] = crm.status
        crm_row_dict["Catalogue File Name"] = catalogue_path
        crm_row_dict["URL"] = crm.url
        # print(crm.chemistry)
        # write ppm info
        for element, method_conc_dict in crm.chemistry_ppm.items():
            element_symbol_only = re.sub(r"\([^)]*\)", "", element)
            if element_symbol_only not in PREFS:
                continue
            for pref_method in PREFS[element_symbol_only]:
                if pref_method not in method_conc_dict:
                    continue
                else:
                    # ADD CONCENTRATION FOR THAT ELEMENT: column title is element name(unit) = concentration value in dict for that method
                    crm_row_dict[str(element)] = method_conc_dict[str(pref_method)]
                    # ADD METHOD TO METHOD COLUMN FOR THAT ELEMENT.
                    crm_row_dict[f"{element_symbol_only} Method"] = pref_method
                    break
                # this will run if no matches are found for the element
                print("no ")

        # write % info
        for element, method_conc_dict in crm.chemistry_wtpercent.items():
            element_symbol_only = re.sub(r"\([^)]*\)", "", element)
            if element_symbol_only not in PREFS:
                continue
            for pref_method in PREFS[element_symbol_only]:
                if pref_method not in method_conc_dict:
                    continue
                else:
                    # ADD CONCENTRATION FOR THAT ELEMENT: column title is element name(unit) = concentration value in dict for that method
                    crm_row_dict[str(element)] = method_conc_dict[str(pref_method)]
                    # ADD METHOD TO METHOD COLUMN FOR THAT ELEMENT.
                    # crm_row_dict[f"{element} Method"] = pref_method
                    break

        # print(crm_row_dict)
        new_row_df = pd.DataFrame(data=crm_row_dict, index=[0])
        all_data_df = pd.concat([all_data_df, new_row_df])

    # TO CSV
    # all_data_df.to_csv("output_processed.csv", index=False, float_format="%f")
    # print(
    #     f"Processed data output to CSV. rows={all_data_df.shape[0]}, cols={all_data_df.shape[1]}"
    # )

    # GENERATE DATAFRAMES FOR PPM AND % VALS ONLY
    ppm_data_df_selected_cols = all_data_df.columns[:1].tolist() + [
        col for col in all_data_df.columns if "(PPM)" in col
    ]
    ppm_data_df = all_data_df[ppm_data_df_selected_cols]
    percent_data_df_selected_cols = all_data_df.columns[:1].tolist() + [
        col for col in all_data_df.columns if "(%)" in col
    ]
    percent_data_df = all_data_df[percent_data_df_selected_cols]
    # TO EXCEL
    # Create an ExcelWriter object
    outputxlsxname = "output.xlsx"
    with pd.ExcelWriter(outputxlsxname, engine="openpyxl") as xlwriter:
        all_data_df.to_excel(xlwriter, sheet_name="All Data", index=False)
        ppm_data_df.to_excel(xlwriter, sheet_name="PPM Data", index=False)
        percent_data_df.to_excel(xlwriter, sheet_name="Percent Data", index=False)

        # Access the workbook and sheet
        xlworkbook = xlwriter.book
        xlworksheet_all = xlwriter.sheets["All Data"]
        xlworksheet_ppm = xlwriter.sheets["PPM Data"]
        xlworksheet_percent = xlwriter.sheets["Percent Data"]

        # Iterate through the cells to find and format the ones with 'INDICATIVE' text, FOR EACH PAGE IN WORKBOOK
        print(f'Processing and Formatting {outputxlsxname} - "All Data" sheet...')
        for row in xlworksheet_all.iter_rows(
            min_row=2, max_col=all_data_df.shape[1], max_row=xlworksheet_all.max_row
        ):
            for cell in row:
                if " INDICATIVE" in str(cell.value):
                    cell.font = Font(strikethrough=True)
                    cell.value = float(cell.value.replace(" INDICATIVE", "").strip())

        print(f'"All Data" sheet formatted.')

        print(f'Processing and Formatting {outputxlsxname} - "PPM Data" sheet...')
        for row in xlworksheet_ppm.iter_rows(
            min_row=2, max_col=ppm_data_df.shape[1], max_row=xlworksheet_ppm.max_row
        ):
            for cell in row:
                if " INDICATIVE" in str(cell.value):
                    cell.font = Font(strikethrough=True)
                    cell.value = float(cell.value.replace(" INDICATIVE", "").strip())
                elif cell.value is None or cell.value == "":
                    cell.value = "T"
        print(f'"PPM Data" sheet formatted.')

        print(f'Processing and Formatting {outputxlsxname} - "Percent Data" sheet...')
        for row in xlworksheet_percent.iter_rows(
            min_row=2,
            max_col=percent_data_df.shape[1],
            max_row=xlworksheet_percent.max_row,
        ):
            for cell in row:
                if " INDICATIVE" in str(cell.value):
                    cell.font = Font(strikethrough=True)
                    cell.value = float(cell.value.replace(" INDICATIVE", "").strip())
                elif cell.value is None or cell.value == "":
                    cell.value = "T"
        print(f'"PPM Data" sheet formatted.')

        "Finishing up..."
    print("Done.")

    # Save the Excel file
    # xlworkbook.save()

    # # output data to txt for testing
    # with open('output.txt',mode='w') as f:
    #     for crm in crms:
    #         f.write(f'start of {crm.id} chemistry:\n')
    #         f.write(json.dumps(crm.chemistry))
    #         f.write(f'\nend of {crm.id} chemistry.\n\n\n')

    # # find list of unique entries in given col
    # unique_methods_list = set(cat_df['Analysis Method'].tolist())
    # methods_counter = Counter(cat_df['Analysis Method'].tolist())
    # print(unique_methods_list)
    # print(f'NUMBER OF UNIQUE METHODS: {len(unique_methods_list)}')
    # print(methods_counter)

    # unique_compound_list = set(catalogue_df['Element'].tolist())
    # print(unique_compound_list)
    # print(f'NUMBER OF UNIQUE COMPOUNDS: {len(unique_compound_list)}')
    # unique_compound_formulas_list_all = []
    # for compound in unique_compound_list:
    #     if ', ' in compound:
    #         unique_compound_formulas_list_all.append(compound.split(', ')[1])
    #         stoich_dict = chemparse.parse_formula(compound.split(', ')[1])
    #         print(stoich_dict)
    # unique_compound_formulas = set(unique_compound_formulas_list_all)
    # print(unique_compound_formulas)
    # print(f'NUMBER OF UNIQUE COMPOUND FORMULAS: {len(unique_compound_formulas)}')

    # for i in catalogue_df.index:


if __name__ == "__main__":
    main()
